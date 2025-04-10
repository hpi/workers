import subprocess
import time
import argparse
from wakeonlan import send_magic_packet
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from datetime import date, datetime, timedelta
import cv2
import numpy as np
from pdf2image import convert_from_path
import pytesseract
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, NamedStyle, Font, PatternFill

def insert_tasks(pxyl_workbook, tasks):
  sheet = pxyl_workbook.active

  task_boxes = []
  start_column = 'C'
  end_column = 'K'
  start_row = 28
  task_index = 0

  for row in range(start_row, 50, 2):
    cell = sheet[f'{start_column}{row}']
    if cell.value is None:
      if task_index >= len(tasks):
        continue

      cell.value = f'  {tasks[task_index]}'
      cell.font = Font(size=16, name='Arial')
      task_index += 1

  return task_boxes

def load_sheet(sheet_path):
  return load_workbook(sheet_path)

def persist_sheet(workbook, sheet_name):
    ws = workbook.active
    ws.print_area = 'A1:M55'  # Adjust the column range 'A:Z' as needed
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    no_border = Border(left=Side(style=None),
                       right=Side(style=None),
                       top=Side(style=None),
                       bottom=Side(style=None))

    # Define a dark gray border style
    gray_border = Border(left=Side(style='thin', color="808080"),
                         right=Side(style='thin', color="808080"),
                         top=Side(style='thin', color="808080"),
                         bottom=Side(style='thin', color="808080"))

    # Apply no border to all cells
    for row in ws.iter_rows(min_row=1, max_row=58, min_col=1, max_col=20):
        for cell in row:
            cell.border = no_border

    if sheet_name == 'sod':
      # Apply dark gray border to specific ranges
      for row in ws.iter_rows(min_row=28, max_row=51, min_col=2, max_col=12):
          for cell in row:
              cell.border = gray_border

      for row in ws.iter_rows(min_row=8, max_row=25, min_col=2, max_col=12):
          for cell in row:
              cell.border = gray_border
    elif sheet_name == 'eod':
      for row in ws.iter_rows(min_row=7, max_row=38, min_col=2, max_col=11):
          for cell in row:
              cell.border = gray_border

    # Save the modified workbook
    workbook.save(f'{sheet_name}-modified.xlsx')

def convert_to_pdf(input_path):
  subprocess.run(['libreoffice', '--headless', '--convert-to', 'pdf', input_path])

def insert_date(pxyl_workbook):
  sheet = pxyl_workbook.active

  sheet['C6'].font = Font(size=16, name='Arial')
  sheet['C6'].value = f'  {str(date.today())}'

def download_google_sheet_as_xlsx(file_id, credentials_path, output_path):
    scopes = ['https://www.googleapis.com/auth/spreadsheets.readonly', 'https://www.googleapis.com/auth/drive.readonly']
    credentials = Credentials.from_service_account_file(credentials_path, scopes=scopes)
    drive_service = build('drive', 'v3', credentials=credentials)
    request = drive_service.files().export_media(fileId=file_id, mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response = request.execute()

    with open(output_path, 'wb') as f:
        f.write(response)

def print_pdf(printer_name, pdf_path):
    subprocess.run(["lpr", "-P", printer_name, pdf_path])

def get_tasks():
    # Execute the command and capture the output
    result = subprocess.run(['task', 'export', 'next'], stdout=subprocess.PIPE)
    tasks_json = result.stdout.decode('utf-8')

    # Parse the JSON data
    try:
        tasks = json.loads(tasks_json)
    except json.JSONDecodeError:
        return ["Error: Could not parse tasks"]

    # Calculate the current datetime and 24 hours from now
    now = datetime.now()
    in_24_hours = now + timedelta(days=1)

    # Filter tasks due within the next 24 hours or overdue
    tasks_due_soon_or_overdue = [
        f"{task['project'] + ': ' if 'project' in task else ''}{task['description']} ({task['id']})"
        for task in tasks
        if 'due' in task and datetime.strptime(task['due'], "%Y%m%dT%H%M%SZ") <= in_24_hours
        and (task['status'] != 'completed' or datetime.strptime(task['due'], "%Y%m%dT%H%M%SZ") < now)
    ]

    return tasks_due_soon_or_overdue

def main():
    parser = argparse.ArgumentParser(description="Download Google Sheet and print it.")
    parser.add_argument("sheet_id", help="Sheet ID to download")
    parser.add_argument("type", help="Type of BTF document to print")
    parser.add_argument("printer_name", help="Name of the printer")

    args = parser.parse_args()

    # Configuration
    # Canon MG3600
    #mac_address_of_printer = '74:BF:C0:AF:C0:DD'
    # TR47
    mac_address_of_printer = 'DC:C2:C9:0A:17:BF'
    printer_name = args.printer_name
    sheet_id = args.sheet_id
    btf_type = args.type
    output_path = f'{btf_type}.pdf'
    credentials_path = '/home/mads/.google-credentials.json'

    # Step 1: Wake up the printer
    send_magic_packet(mac_address_of_printer)

    # Step 2: Give the printer some time to wake up
    time.sleep(20)

    sheet_name = f'{btf_type}.xlsx'
    download_google_sheet_as_xlsx(sheet_id, credentials_path, sheet_name)
    workbook = load_sheet(sheet_name)

    insert_date(workbook)

    if btf_type == 'sod':
      task_items = get_tasks()

      insert_tasks(workbook, task_items)

    persist_sheet(workbook, btf_type)
    convert_to_pdf(f'{btf_type}-modified.xlsx')

    # Step 4: Print the PDF
    print_pdf(printer_name, f'{btf_type}-modified.pdf')

if __name__ == '__main__':
    main()

